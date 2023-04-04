import openpyxl


def getExcelOneRowData(docName, sheetNum, rowNum):
    wb = openpyxl.load_workbook(docName)
    # 获取某一个页签对象
    sheet = wb[wb.sheetnames[sheetNum]]
    col=1
    res = list()
    while sheet.cell(rowNum, col).value is not None:
        res.append(sheet.cell(rowNum, col).value)
        col += 1
    return res
